"""
importer.py — Import test cases from Capstone_Final.xlsx into eval.duckdb.
Safe to call multiple times; skips import if test_cases table already has data.
"""

import os
import openpyxl

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "data", "Capstone_Final.xlsx")

DEFAULT_TIME_THRESHOLD_MS = 10000.0  # 10 seconds generous default


def import_from_excel(db) -> dict:
    """
    Read Capstone_Final.xlsx and insert test cases into the database.

    Args:
        db: the eval_db module (passed in to avoid circular import issues)

    Returns:
        dict with counts of inserted records per category.
    """
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"Excel file not found: {EXCEL_PATH}")

    wb = openpyxl.load_workbook(EXCEL_PATH)
    counts = {"conversational": 0, "sql": 0, "performance": 0}

    # ── Conversational ────────────────────────────────────────────────────────
    ws = wb["Conversational"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        q_no, question, q_type, eval_purpose, expected_output, *_ = list(row) + [None] * 8
        if not question:
            continue
        db.add_test_case(
            category="conversational",
            question=str(question).strip(),
            expected_output=str(expected_output).strip() if expected_output else None,
        )
        counts["conversational"] += 1

    # ── SQL ───────────────────────────────────────────────────────────────────
    ws = wb["SQL"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        question, golden_sql, *_ = list(row) + [None] * 4
        if not question:
            continue
        db.add_test_case(
            category="sql",
            question=str(question).strip(),
            golden_sql=str(golden_sql).strip() if golden_sql else None,
        )
        counts["sql"] += 1

    # ── Performance ───────────────────────────────────────────────────────────
    ws = wb["Performance"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        question, _output_text, time_threshold_ms, expected_rows, *_ = list(row) + [None] * 6
        if not question:
            continue

        db.add_test_case(
            category="performance",
            question=str(question).strip(),
            expected_rows=int(expected_rows) if expected_rows is not None else None,
            time_threshold_ms=float(time_threshold_ms) if time_threshold_ms is not None else DEFAULT_TIME_THRESHOLD_MS,
        )
        counts["performance"] += 1

    return counts
